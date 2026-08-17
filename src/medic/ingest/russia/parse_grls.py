"""Parse the GRLS bulk export (``background/grls.zip``) into drug records.

The export is a zip of 8 ``.xlsx`` register files, one per registration state.
The member filenames are Cyrillic; on many terminals the zip's cp437 storage
garbles them, so we iterate ``ZipFile.infolist()`` **by index** and decode the
names defensively rather than looking anything up by name.

Every register shares the same 17-column layout. A banner occupies the first
rows; the real header is the row containing ``Дата регистрации`` and data begins
two rows below it. Column indices (0-based) used here:

======  =====================================================  ===================================
Index   Cyrillic header                                        English meaning
======  =====================================================  ===================================
2       Номер регистрационного удостоверения                   Registration certificate number
3       Дата регистрации                                       Registration date (DD.MM.YYYY)
6       Юридическое лицо ...                                   Marketing-authorisation holder
8       Торговое наименование лекарственного препарата         Trade name
9       Международное непатентованное или химическое ...       INN or chemical name (МНН)
======  =====================================================  ===================================

There is **no indication column** in the GRLS export, so Russia contributes a
drug list only — no indications (matching prior behaviour).

Register scope
--------------
We include every register that represents a *currently valid* registration and
exclude the two that represent former registrations:

  - included: Действующий (active), Изменённый (modified), Выдано по правилам
    ЕАЭС (EAEU-issued), Действует на подтверждении ... (active, pending
    confirmation), Приостановлено применение (suspended), Действует в
    иностранных упаковках (active in foreign packaging).
  - excluded: Исключённый (excluded/struck-off) and Истёкший (expired).
"""

from __future__ import annotations

import logging
import tempfile
import zipfile
from pathlib import Path

import openpyxl

from medic.mention import assign_mention

logger = logging.getLogger(__name__)

# Column indices in the register sheets (0-based).
COL_REG_NUMBER = 2
COL_REG_DATE = 3
COL_HOLDER = 6
COL_TRADE_NAME = 8
COL_INN = 9

# Cyrillic fragment identifying the header row.
_HEADER_MARKER = "Дата регистрации"

# Register-state name fragments (in the member filename, after cp866 decoding)
# that we EXCLUDE because they are former, not current, registrations.
_EXCLUDED_REGISTER_FRAGMENTS = ("Исключённый", "Истёкший")

# Placeholder values GRLS uses when there is no meaningful INN.
_EMPTY_INN = {"", "~", "-", "nan", "none"}


def _decode_member_name(raw_name: str) -> str:
    """Best-effort decode of a (possibly cp437-garbled) Cyrillic zip member name."""
    try:
        return raw_name.encode("cp437").decode("cp866")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return raw_name


def _reg_date_to_iso(value) -> str:
    """Convert a GRLS registration date (``DD.MM.YYYY``) to ``YYYYMMDD``.

    Returns an empty string if the value is missing or unparseable. openpyxl may
    hand us a ``datetime`` for date-typed cells or a raw string; handle both.
    """
    if value is None:
        return ""
    # openpyxl may return a datetime for date-typed cells.
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y%m%d")
        except (ValueError, AttributeError):
            return ""
    s = str(value).strip()
    if not s:
        return ""
    parts = s.split(".")
    if len(parts) == 3 and all(p.isdigit() for p in parts):
        day, month, year = parts
        if len(year) == 4:
            return f"{year}{int(month):02d}{int(day):02d}"
    return ""


def _find_header_row(ws) -> int:
    """Return the 1-based row index of the header row, or -1 if not found."""
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        if any(cell and _HEADER_MARKER in str(cell) for cell in row):
            return ri
    return -1


def _parse_sheet(ws, register_label: str) -> list[dict]:
    """Extract raw drug rows from one register sheet."""
    header_row = _find_header_row(ws)
    if header_row < 0:
        logger.warning("No header row found in register %r; skipping", register_label)
        return []
    data_start = header_row + 1
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        # A register row must have at least COL_INN columns.
        if len(row) <= COL_INN:
            continue
        inn = row[COL_INN]
        trade = row[COL_TRADE_NAME]
        # Fully empty spacer row.
        if inn is None and trade is None:
            continue
        inn_s = str(inn).strip() if inn is not None else ""
        trade_s = str(trade).strip() if trade is not None else ""
        # Prefer the INN; fall back to the trade name when the INN is a
        # GRLS placeholder (herbals / complex products carry no INN).
        if inn_s.lower() in _EMPTY_INN:
            name = trade_s
            name_source = "trade_name"
        else:
            name = inn_s
            name_source = "inn"
        if not name:
            continue
        rows.append(
            {
                "name": name,
                "name_source": name_source,
                "trade_name": trade_s,
                "reg_number": str(row[COL_REG_NUMBER]).strip() if row[COL_REG_NUMBER] else "",
                "reg_date": _reg_date_to_iso(row[COL_REG_DATE]),
                "holder": str(row[COL_HOLDER]).strip() if row[COL_HOLDER] else "",
                "register": register_label,
            }
        )
    return rows


def parse_grls_zip(zip_path: Path) -> list[dict]:
    """Parse all currently-valid registers in the GRLS zip into drug records.

    Deduplicates by drug name (the Cyrillic INN, or trade-name fallback),
    keeping the earliest registration date and recording the set of
    registration numbers seen.

    Returns a list of records shaped for ``ground_records``: each carries a
    ``source: "RUSSIA"``, ``source_name`` (the name to ground) and provenance
    fields (``original_name_ru``, ``approval_date``, ``application_number``,
    ``trade_name``).
    """
    aggregated: dict[str, dict] = {}
    total_rows = 0

    with zipfile.ZipFile(zip_path) as zf:
        for index, info in enumerate(zf.infolist()):
            member_name = _decode_member_name(info.filename)
            if not member_name.lower().endswith(".xlsx"):
                continue
            if any(frag in member_name for frag in _EXCLUDED_REGISTER_FRAGMENTS):
                logger.info("Skipping former-registration register: %s", member_name)
                continue
            register_label = member_name.rsplit("-1-", 1)[-1].removesuffix(".xlsx")
            logger.info("Parsing register [%d]: %s", index, register_label)
            # Extract to a temp file (openpyxl needs a seekable file/path, and
            # some Cyrillic-named members are large — stream via the index).
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(zf.read(info))
                tmp_path = Path(tmp.name)
            try:
                wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                for ws in wb.worksheets:
                    rows = _parse_sheet(ws, register_label)
                    total_rows += len(rows)
                    for r in rows:
                        key = r["name"]
                        existing = aggregated.get(key)
                        if existing is None:
                            aggregated[key] = {
                                "name": r["name"],
                                "name_source": r["name_source"],
                                "trade_name": r["trade_name"],
                                "reg_date": r["reg_date"],
                                "reg_numbers": {r["reg_number"]} if r["reg_number"] else set(),
                            }
                        else:
                            # Keep the earliest non-empty registration date.
                            if r["reg_date"] and (
                                not existing["reg_date"] or r["reg_date"] < existing["reg_date"]
                            ):
                                existing["reg_date"] = r["reg_date"]
                            if r["reg_number"]:
                                existing["reg_numbers"].add(r["reg_number"])
                wb.close()
            finally:
                tmp_path.unlink(missing_ok=True)

    records: list[dict] = []
    for agg in aggregated.values():
        reg_numbers = sorted(agg["reg_numbers"])
        record = {
            "source": "RUSSIA",
            # ``source_name`` holds the verbatim Cyrillic name. The shared
            # translation stage (DeepL via babelon) translates it to English and
            # overwrites this field before grounding; the Cyrillic is preserved
            # in ``original_name_ru`` and in the Babelon translation row (I-7).
            "source_name": agg["name"],
            "original_name_ru": agg["name"],
            "approval_date": agg["reg_date"],
        }
        if reg_numbers:
            # One drug (INN) can span many registration certificates; expose the
            # first as the primary application number and keep the rest.
            record["application_number"] = reg_numbers[0]
            if len(reg_numbers) > 1:
                record["application_numbers"] = reg_numbers
        if agg["trade_name"] and agg["name_source"] == "inn":
            record["trade_name"] = agg["trade_name"]
        # Mint the MEDICNE id from the verbatim Cyrillic literal at extraction (I-9),
        # before translation overwrites ``source_name``.
        assign_mention(record, "drugs", literal=agg["name"])
        records.append(record)

    logger.info(
        "Parsed %d GRLS rows -> %d unique drug records (by name)",
        total_rows,
        len(records),
    )
    return records
